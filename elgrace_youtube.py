from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook

chrome_options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# 주일예배 ws1
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-zkkGQSRCmNCmPsHTiQUvUXc')
time.sleep(3)

for i in range(0, 50):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

wb = Workbook()

ws1 = wb.create_sheet("주일예배", 0)
ws1.append(['제목'])

ws2 = wb.create_sheet("예배_스트리밍", 1)
ws2.append(['제목'])

ws3 = wb.create_sheet("수요예배", 2)
ws3.append(['제목'])

ws4 = wb.create_sheet("특강", 3)
ws4.append(['제목'])

ws5 = wb.create_sheet("40일_기도회", 4)
ws5.append(['제목'])

ws6 = wb.create_sheet("아침_기도회", 5)
ws6.append(['제목'])

ws7 = wb.create_sheet("2021_청년부_여름수련회_특강", 6)
ws7.append(['제목'])

ws8 = wb.create_sheet("2021_청년부_여름수련회_저녁강의", 7)
ws8.append(['제목'])

ws9 = wb.create_sheet("주은혜_스토리", 8)
ws9.append(['제목'])

wb.remove_sheet(wb['Sheet'])

for title in titles:
    print(title.text)
    ws1.append([title.text])

# 예배_스트리밍 ws2
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-zl6BZ7wC-aVqvIwo7MBCIG9')
time.sleep(3)

for i in range(0, 50):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

for title in titles:
    print(title.text)
    ws2.append([title.text])

# 수요예배 ws3
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-znnod0w6WBIuYl8nWcSe3fK')
time.sleep(3)

for i in range(0, 50):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

for title in titles:
    print(title.text)
    ws3.append([title.text])

# 특강 ws4
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-zlN3SGf2znZnfltmhEyHNc-')
time.sleep(3)

for i in range(0, 10):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

for title in titles:
    print(title.text)
    ws4.append([title.text])

# 40일_기도회 ws5
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-zm2OLKFFkOyIQ6CawfaHRGl')
time.sleep(3)

for i in range(0, 30):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

for title in titles:
    print(title.text)
    ws5.append([title.text])

# 아침_기도회 ws6
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-zmrHwCqRucVPNhHk52DOYwx')
time.sleep(3)

for i in range(0, 10):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

for title in titles:
    print(title.text)
    ws6.append([title.text])

# 2021_청년부_여름수련회_특강 ws7
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-zmeiFV9yOC8SKCDpRy0NO4B')
time.sleep(3)

for i in range(0, 10):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

for title in titles:
    print(title.text)
    ws7.append([title.text])

# 2021_청년부_여름수련회_저녁강의 ws8
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-zkPx2OapLWtFGg2e-BoEDYc')
time.sleep(3)

for i in range(0, 10):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

for title in titles:
    print(title.text)
    ws8.append([title.text])

# 주은혜_스토리 ws9
driver.get(url='https://www.youtube.com/playlist?list=PLmaMHgLSo-zk7V8bPkvZkHrzmyaB8oyXW')
time.sleep(3)

for i in range(0, 10):
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)

titles = driver.find_elements(By.CSS_SELECTOR, '#video-title')
print(titles)

for title in titles:
    print(title.text)
    ws9.append([title.text])

wb.save('C:/Users/DW-PC/Documents/elgrace_youtube/Youtube.xlsx')
wb.close()